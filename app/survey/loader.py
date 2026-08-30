from __future__ import annotations

import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
EXCEL_ERROR_VALUES = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A"}


@dataclass
class SurveyDataset:
    source_name: str
    sheets: list[str]
    frames: dict[str, pd.DataFrame]
    warnings: list[str] = field(default_factory=list)

    @property
    def frame(self) -> pd.DataFrame:
        return next(iter(self.frames.values()))

    @property
    def record_count(self) -> int:
        return int(sum(len(frame) for frame in self.frames.values()))


def resolve_file_input(file_path: str | None = None, file: Any = None) -> Path:
    candidate: str | None = file_path
    if isinstance(file, str):
        candidate = candidate or file
    elif isinstance(file, dict):
        candidate = candidate or file.get("path") or file.get("file_path")
        if not candidate and file.get("url"):
            raise ValueError("El archivo recibido por URL debe descargarse a un almacenamiento controlado antes de procesarlo.")
    if not candidate:
        raise ValueError("Adjunta un archivo Excel o CSV, o proporciona file_path durante las pruebas locales.")

    path = Path(str(candidate)).expanduser().resolve()
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("Formato no compatible. Usa .xlsx, .xls o .csv.")
    if not path.is_file():
        raise FileNotFoundError("No encontré el archivo cargado en el almacenamiento de trabajo.")
    configured_max_bytes = os.getenv("MAX_UPLOAD_BYTES", "").strip()
    try:
        max_bytes = int(configured_max_bytes or str(50 * 1024 * 1024))
    except (TypeError, ValueError):
        max_bytes = 50 * 1024 * 1024
    if path.stat().st_size > max_bytes:
        raise ValueError(f"El archivo supera el límite permitido de {max_bytes // (1024 * 1024)} MB.")
    return path


def _clean_frame(frame: pd.DataFrame) -> pd.DataFrame:
    frame = frame.dropna(axis=0, how="all").dropna(axis=1, how="all").copy()
    names: list[str] = []
    seen: dict[str, int] = {}
    for index, column in enumerate(frame.columns):
        base = str(column).strip() if str(column).strip() else f"columna_{index + 1}"
        seen[base] = seen.get(base, 0) + 1
        names.append(base if seen[base] == 1 else f"{base} ({seen[base]})")
    frame.columns = names
    return frame.reset_index(drop=True)


def _header_row(raw: pd.DataFrame) -> int:
    best_index, best_score = 0, -1
    for index in range(min(len(raw), 20)):
        row = raw.iloc[index]
        nonempty = [str(value).strip() for value in row.tolist() if pd.notna(value) and str(value).strip()]
        text_bonus = sum(1 for value in nonempty if len(value) > 2 or "?" in value)
        score = len(nonempty) + text_bonus * 0.25
        if score > best_score:
            best_index, best_score = index, score
    return best_index


def _read_excel(path: Path) -> tuple[dict[str, pd.DataFrame], list[str]]:
    warnings = _formula_warnings(path)
    raw_book = pd.read_excel(path, sheet_name=None, header=None, dtype=object)
    frames: dict[str, pd.DataFrame] = {}
    for sheet_name, raw in raw_book.items():
        raw = _clean_frame(raw)
        if raw.empty:
            warnings.append(f"La hoja '{sheet_name}' está vacía y fue omitida.")
            continue
        header = _header_row(raw)
        frame = raw.iloc[header + 1 :].copy()
        frame.columns = [str(value).strip() if pd.notna(value) and str(value).strip() else f"columna_{i + 1}" for i, value in enumerate(raw.iloc[header].tolist())]
        frame = _clean_frame(frame)
        error_count = sum(
            1
            for value in frame.to_numpy().flat
            if isinstance(value, str) and value.strip().upper() in EXCEL_ERROR_VALUES
        )
        if error_count:
            frame = frame.replace(list(EXCEL_ERROR_VALUES), pd.NA)
            warnings.append(
                f"La hoja '{sheet_name}' contiene {error_count} error(es) de Excel sin valor calculado utilizable; esas celdas se trataron como vacías."
            )
        frames[str(sheet_name)] = frame
    return frames, warnings


def _formula_warnings(path: Path) -> list[str]:
    """Report formulas without executing workbook content."""
    if path.suffix.lower() != ".xlsx":
        return []
    warnings: list[str] = []
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        for sheet in workbook.worksheets:
            formula_count = sum(
                1
                for row in sheet.iter_rows()
                for cell in row
                if cell.data_type == "f"
                or (isinstance(cell.value, str) and cell.value.startswith("="))
            )
            if formula_count:
                warnings.append(
                    f"La hoja '{sheet.title}' contiene {formula_count} celda(s) con fórmula. No se ejecutaron; sólo se usaron los valores calculados guardados en el archivo."
                )
    except (OSError, ValueError, TypeError):
        # pandas/openpyxl will still surface the actual read error if the file is invalid.
        return warnings
    finally:
        if workbook is not None:
            workbook.close()
    return warnings


def _read_csv(path: Path) -> tuple[dict[str, pd.DataFrame], list[str]]:
    errors: list[str] = []
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            raw = pd.read_csv(path, header=None, dtype=object, encoding=encoding)
            raw = _clean_frame(raw)
            header = _header_row(raw)
            frame = raw.iloc[header + 1 :].copy()
            frame.columns = [str(value).strip() if pd.notna(value) and str(value).strip() else f"columna_{i + 1}" for i, value in enumerate(raw.iloc[header].tolist())]
            return {"CSV": _clean_frame(frame)}, errors
        except UnicodeDecodeError as exc:
            errors.append(str(exc))
    raise ValueError("No pude leer la codificación del CSV.")


def load_dataset(file_path: str | None = None, file: Any = None, sheet_name: str | None = None) -> SurveyDataset:
    path = resolve_file_input(file_path, file)
    if path.suffix.lower() == ".csv":
        frames, warnings = _read_csv(path)
    else:
        frames, warnings = _read_excel(path)
    if not frames:
        raise ValueError("No encontré hojas con registros utilizables en el archivo.")
    if sheet_name:
        if sheet_name not in frames:
            available = ", ".join(frames)
            raise ValueError(f"No encontré la hoja '{sheet_name}'. Hojas disponibles: {available}.")
        frames = {sheet_name: frames[sheet_name]}
    return SurveyDataset(path.name, list(frames), frames, warnings)
